import os
from openpyxl import load_workbook
import pandas as pd
from colorama import Fore

def load_file(path):
    wb = None 

    if path.endswith(('.xls', '.xlsx')):
        if os.path.exists(path):
            if os.path.isfile(path):
                try:
                    if path.endswith('.xls'):
                        xlsx_path = path.replace('.xls', '.xlsx')
                        df = pd.read_excel(path)
                        df.to_excel(xlsx_path, index=False) 
                        print(Fore.GREEN + '✔' + Fore.WHITE + " | The .xls file has been converted to .xlsx successfully!")
                        path = xlsx_path

                    wb = load_workbook(path)
                    print(Fore.GREEN + '✔' + Fore.WHITE + " | The file has been successfully uploaded!")
                    print(Fore.LIGHTBLUE_EX + '→' + Fore.WHITE + ' | Found workbooks on file: ' + ", ".join(map(str, wb.sheetnames)))
                except Exception as e:
                    print(Fore.RED + '✘' + Fore.WHITE + " | An error occurred when uploading file: " + Fore.RED + f"{e}")
            else:
                print(Fore.RED + '✘' + Fore.WHITE + " | The path entered is not a file!")
        else:
            print(Fore.RED + '✘' + Fore.WHITE + " | The entered path is invalid!")
    else:
        print(Fore.RED + '✘' + Fore.WHITE + " | This file format " + Fore.RED + "is not supported" + Fore.WHITE + "!" + Fore.WHITE + " Use " + Fore.GREEN + ".xls " + Fore.WHITE + "or " + Fore.GREEN + ".xlsx" + Fore.WHITE + ".")

    return wb